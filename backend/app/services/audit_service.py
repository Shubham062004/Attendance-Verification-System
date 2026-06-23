import csv
import io
from datetime import datetime, time, UTC
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from sqlalchemy import desc, or_
from sqlalchemy.orm import Session
from app.models.audit import AuditLog


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#475569"))
        
        # Draw header
        self.drawString(inch, 10.4 * inch, "Smart Attendance Verification System")
        self.drawRightString(7.5 * inch, 10.4 * inch, "Audit Logs Report")
        self.setStrokeColor(colors.HexColor("#e2e8f0"))
        self.setLineWidth(0.5)
        self.line(inch, 10.3 * inch, 7.5 * inch, 10.3 * inch)

        # Draw footer
        self.setFont("Helvetica", 8)
        self.drawString(inch, 0.4 * inch, f"Generated on {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S')} UTC")
        self.drawRightString(7.5 * inch, 0.4 * inch, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()


class AuditService:
    @staticmethod
    def log_action(
        db: Session,
        actor_id: int | None,
        actor_name: str | None,
        actor_role: str | None,
        action_type: str,
        entity_type: str | None,
        entity_id: int | None,
        description: str,
        metadata_json: dict[str, Any] | None = None,
        ip_address: str | None = None,
        user_agent: str | None = None,
    ) -> AuditLog:
        log_entry = AuditLog(
            actor_id=actor_id,
            actor_name=actor_name,
            actor_role=actor_role,
            action_type=action_type,
            entity_type=entity_type,
            entity_id=entity_id,
            description=description,
            metadata_json=metadata_json,
            ip_address=ip_address,
            user_agent=user_agent,
        )
        db.add(log_entry)
        db.commit()
        db.refresh(log_entry)
        return log_entry

    @staticmethod
    def get_by_id(db: Session, log_id: int) -> AuditLog | None:
        return db.query(AuditLog).filter(AuditLog.id == log_id).first()

    @staticmethod
    def get_by_user(db: Session, user_id: int) -> list[AuditLog]:
        return db.query(AuditLog).filter(AuditLog.actor_id == user_id).order_by(desc(AuditLog.created_at)).all()

    @staticmethod
    def get_by_entity(db: Session, entity_id: int, entity_type: str | None = None) -> list[AuditLog]:
        query = db.query(AuditLog).filter(AuditLog.entity_id == entity_id)
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        return query.order_by(desc(AuditLog.created_at)).all()

    @staticmethod
    def _build_filter_query(
        db: Session,
        action_type: str | None = None,
        actor_role: str | None = None,
        entity_type: str | None = None,
        actor_id: int | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        search: str | None = None,
    ):
        query = db.query(AuditLog)

        if action_type:
            query = query.filter(AuditLog.action_type == action_type)
        if actor_role:
            query = query.filter(AuditLog.actor_role == actor_role)
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        if actor_id is not None:
            query = query.filter(AuditLog.actor_id == actor_id)

        if start_date:
            try:
                start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d").date(), time.min)
                query = query.filter(AuditLog.created_at >= start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d").date(), time.max)
                query = query.filter(AuditLog.created_at <= end_dt)
            except ValueError:
                pass

        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    AuditLog.description.ilike(search_pattern),
                    AuditLog.actor_name.ilike(search_pattern),
                    AuditLog.action_type.ilike(search_pattern),
                    AuditLog.entity_type.ilike(search_pattern),
                )
            )

        return query

    @staticmethod
    def search_logs(
        db: Session,
        page: int = 1,
        size: int = 50,
        action_type: str | None = None,
        actor_role: str | None = None,
        entity_type: str | None = None,
        actor_id: int | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        search: str | None = None,
    ) -> tuple[list[AuditLog], int]:
        query = AuditService._build_filter_query(
            db, action_type, actor_role, entity_type, actor_id, start_date, end_date, search
        )
        total = query.count()
        items = query.order_by(desc(AuditLog.created_at)).offset((page - 1) * size).limit(size).all()
        return items, total

    @staticmethod
    def export_logs(
        db: Session,
        format_type: str,
        action_type: str | None = None,
        actor_role: str | None = None,
        entity_type: str | None = None,
        actor_id: int | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        search: str | None = None,
    ) -> bytes:
        query = AuditService._build_filter_query(
            db, action_type, actor_role, entity_type, actor_id, start_date, end_date, search
        )
        logs = query.order_by(desc(AuditLog.created_at)).all()

        if format_type.lower() == "csv":
            return AuditService._generate_csv(logs)
        elif format_type.lower() == "excel":
            return AuditService._generate_excel(logs)
        elif format_type.lower() == "pdf":
            return AuditService._generate_pdf(logs)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")

    @staticmethod
    def _generate_csv(logs: list[AuditLog]) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Timestamp (UTC)",
            "Actor Name",
            "Actor Role",
            "Action Type",
            "Entity Type",
            "Entity ID",
            "Description",
            "IP Address",
            "User Agent",
        ])
        for log in logs:
            writer.writerow([
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
                log.actor_name or "System",
                log.actor_role or "System",
                log.action_type,
                log.entity_type or "—",
                log.entity_id if log.entity_id is not None else "—",
                log.description,
                log.ip_address or "—",
                log.user_agent or "—",
            ])
        return output.getvalue().encode("utf-8")

    @staticmethod
    def _generate_excel(logs: list[AuditLog]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        headers = [
            "Timestamp (UTC)",
            "Actor Name",
            "Actor Role",
            "Action Type",
            "Entity Type",
            "Entity ID",
            "Description",
            "IP Address",
            "User Agent",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for log in logs:
            ws.append([
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
                log.actor_name or "System",
                log.actor_role or "System",
                log.action_type,
                log.entity_type or "—",
                log.entity_id if log.entity_id is not None else "—",
                log.description,
                log.ip_address or "—",
                log.user_agent or "—",
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 1:
                    cell.font = Font(name="Calibri", size=10)
                    if cell.column in (1, 3, 5, 6, 8):
                        cell.alignment = center_align
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @staticmethod
    def _generate_pdf(logs: list[AuditLog]) -> bytes:
        stream = io.BytesIO()
        doc = SimpleDocTemplate(
            stream,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=54,
            bottomMargin=54
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name="ReportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1e1b4b"),
            alignment=0,
            spaceAfter=15
        ))
        styles.add(ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            textColor=colors.white,
            alignment=1
        ))
        styles.add(ParagraphStyle(
            name="TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            alignment=1
        ))
        styles.add(ParagraphStyle(
            name="TableCellLeft",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            alignment=0
        ))

        story = []
        story.append(Spacer(1, 10))
        story.append(Paragraph("System Audit Trail Report", styles["ReportTitle"]))
        story.append(Spacer(1, 10))

        table_content = [[
            Paragraph("Timestamp (UTC)", styles["TableHeader"]),
            Paragraph("Actor", styles["TableHeader"]),
            Paragraph("Role", styles["TableHeader"]),
            Paragraph("Action Type", styles["TableHeader"]),
            Paragraph("Entity", styles["TableHeader"]),
            Paragraph("Description", styles["TableHeader"])
        ]]

        for log in logs:
            table_content.append([
                Paragraph(log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "", styles["TableCell"]),
                Paragraph(log.actor_name or "System", styles["TableCell"]),
                Paragraph(log.actor_role or "System", styles["TableCell"]),
                Paragraph(log.action_type, styles["TableCell"]),
                Paragraph(f"{log.entity_type or '—'} ({log.entity_id if log.entity_id is not None else '—'})", styles["TableCell"]),
                Paragraph(log.description, styles["TableCellLeft"])
            ])

        col_widths = [90, 80, 50, 90, 80, 150]

        t = Table(table_content, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f8fafc"), colors.white]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))

        story.append(t)
        doc.build(story, canvasmaker=NumberedCanvas)
        return stream.getvalue()
